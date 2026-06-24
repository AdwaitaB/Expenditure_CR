

import streamlit as st
import pandas as pd
import plotly.express as px
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# ─── Page Config ──────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="AB Group – Expense Analyser",
    page_icon="📊",
    layout="wide",
)

# ─── CSS: fix white-on-white tab text + general styling ───────────────────────
st.markdown("""
<style>
    /* Tab text color fix */
    .stTabs [data-baseweb="tab"] { color: #1E3A5F !important; font-weight: 600; }
    .stTabs [data-baseweb="tab"][aria-selected="true"] { color: #2563EB !important; border-bottom: 3px solid #2563EB; }
    .stTabs [data-baseweb="tab-list"] { background: #F0F4FF; border-radius: 10px; padding: 4px; gap: 4px; }

    /* Header */
    .main-header {
        background: linear-gradient(135deg, #1E3A5F 0%, #2563EB 100%);
        padding: 1.2rem 2rem; border-radius: 12px;
        margin-bottom: 1.5rem; color: white;
    }

    /* Metrics */
    div[data-testid="stMetric"] {
        background: white; border-radius: 10px;
        padding: 0.8rem 1rem;
        box-shadow: 0 1px 4px rgba(0,0,0,0.08);
    }
    div[data-testid="stMetricValue"] { color: #1E3A5F !important; }
    div[data-testid="stMetricLabel"] { color: ##073482 !important; }
    div[data-testid="stMetricLabel"] *,
    div[data-testid="stMetricLabel"] p,
    div[data-testid="stMetricLabel"] label {
        color: #111827 !important;
}

    /* Buttons */
    .stButton > button { border-radius: 8px; font-weight: 600; }

    footer { visibility: hidden; }
</style>
""", unsafe_allow_html=True)

# ─── Category → Main Category Map ─────────────────────────────────────────────
# Based directly on the summary file structure
SUMMARY_STRUCTURE = {
    "Kitchen Consumables (a)": [
        "Grocery", "Vegetable", "Fruit", "Milk & Products",
        "Nonveg", "Colddrink", "Icecream",
    ],
    "Consumables (b)": [
        "Gas", "Housekeeping", "Decoration",
    ],
    "Monthly Expenses ( C)": [
        "Staff Salary", "PSTL Salary", "Electricity", "Fuel", "PSTL-Fuel",
        "Internet", "Mobile Exps", "Generator Rent", "Uniform & Shoes Expenses",
        "Printing and Stationery", "Bank Charges", "Packing", "Medical",
        "Paint", "Transport", "Advertising Expenses", "PSTL-Advertising Expenses",
        "Material", "Professional Fees", "Anciliary Services",
        "Repairs Civil", "Repairs Furniture", "Repairs Computer",
        "Repairs Electrical", "Labour Charges", "Conveyence",
    ],
    "Capital Expenses (d)": [
        "Hardware", "Electrical", "Civil", "Gardening", "Utensils",
        "Sports Material", "Furniture", "Cloth & Mattress", "Machinery",
        "Fabrication", "VIP Guest Expenses", "Additional PSTL",
        "PSTL-Sports Material", "PSTL-Furniture", "PSTL-AC",
    ],
}

# Flat map: expense_head → main_category_group
HEAD_TO_GROUP = {}
for group, heads in SUMMARY_STRUCTURE.items():
    for h in heads:
        HEAD_TO_GROUP[h.strip().lower()] = group

ALL_KNOWN_HEADS = sorted([h for heads in SUMMARY_STRUCTURE.values() for h in heads])


def get_group(expense_head):
    if not expense_head:
        return "Other"
    return HEAD_TO_GROUP.get(str(expense_head).strip().lower(), "Other")


def fmt_inr(n):
    try:
        n = float(n)
        if n >= 1e7:   return f"₹{n/1e7:.2f}Cr"
        if n >= 1e5:   return f"₹{n/1e5:.2f}L"
        if n >= 1000:  return f"₹{n/1000:.1f}K"
        return f"₹{n:,.0f}"
    except:
        return "₹0"


# ─── Parser ───────────────────────────────────────────────────────────────────
def parse_expense_sheet(file_obj):
    """
    Parse the new-format expense Excel.
    F column (index 5) = Expense Head — use directly.
    Returns (DataFrame of records, sheet_name, month_label)
    """
    xl = pd.ExcelFile(file_obj)
    all_records = []
    sheet_name = xl.sheet_names[0]  # use first sheet

    df_raw = pd.read_excel(xl, sheet_name=sheet_name, header=None)

    # Find header row (row with "Expense Head" in col 5)
    header_idx = 0
    for i in range(min(5, len(df_raw))):
        val = str(df_raw.iloc[i, 5]).strip()
        if "Expense Head" in val or "expense head" in val.lower():
            header_idx = i
            break

    data = df_raw.iloc[header_idx + 1:].copy()
    data.columns = range(len(data.columns))

    # Col indices (0-based): 0=Statement, 1=Date, 2=Description, 3=Debit, 4=Credit, 5=ExpenseHead, 6=Bills
    for _, row in data.iterrows():
        # Skip total/empty rows
        stmt = str(row.get(0, "")).strip()
        if stmt.lower() in ("total", "nan", "") and str(row.get(3, "")).strip() in ("nan", ""):
            continue

        raw_debit = str(row.get(3, "")).replace(",", "").strip()
        try:
            debit = float(raw_debit)
        except ValueError:
            continue
        if debit <= 0:
            continue

        expense_head = str(row.get(5, "")).strip()
        if expense_head in ("nan", "Expense Head", ""):
            expense_head = ""

        date_raw = row.get(1, "")
        try:
            if isinstance(date_raw, (pd.Timestamp, datetime)):
                date_str = pd.Timestamp(date_raw).strftime("%d/%m/%y")
            else:
                date_str = str(date_raw).strip().split(" ")[0]
                if "00:00:00" in str(date_raw):
                    date_str = pd.Timestamp(date_raw).strftime("%d/%m/%y")
        except:
            date_str = str(date_raw)

        desc = str(row.get(2, "")).strip()
        account = stmt if stmt not in ("nan", "") else "UPI"
        bills = str(row.get(6, "")).strip()
        bills = "" if bills == "nan" else bills

        all_records.append({
            "Date": date_str,
            "Description": desc[:80],
            "Amount": debit,
            "Expense Head": expense_head,
            "Main Group": get_group(expense_head) if expense_head else "Other",
            "Account": account,
            "Bills": bills,
            "Needs Category": expense_head == "",
        })

    df = pd.DataFrame(all_records)

    # Derive month label from sheet name
    month_label = sheet_name.strip()

    return df, sheet_name, month_label


# ─── Summary Excel Generator ──────────────────────────────────────────────────
def generate_summary_excel(df, month_label):
    """
    Generate summary Excel matching the summary_feb_2026 format exactly.
    Returns BytesIO of the xlsx file.
    """
    # Aggregate by expense head
    totals = {}
    for head in ALL_KNOWN_HEADS:
        mask = df["Expense Head"].str.strip().str.lower() == head.strip().lower()
        totals[head] = round(df.loc[mask, "Amount"].sum(), 2)

    # Catch "Other" entries
    other_mask = df["Main Group"] == "Other"
    other_total = round(df.loc[other_mask, "Amount"].sum(), 2)

    # Compute group subtotals
    kitchen_heads = SUMMARY_STRUCTURE["Kitchen Consumables (a)"]
    consumable_heads = SUMMARY_STRUCTURE["Consumables (b)"]
    monthly_heads = SUMMARY_STRUCTURE["Monthly Expenses ( C)"]
    capital_heads = SUMMARY_STRUCTURE["Capital Expenses (d)"]

    kitchen_total  = sum(totals.get(h, 0) for h in kitchen_heads)
    consumable_total = sum(totals.get(h, 0) for h in consumable_heads)
    monthly_total  = sum(totals.get(h, 0) for h in monthly_heads)
    capital_total  = sum(totals.get(h, 0) for h in capital_heads)
    grand_total    = kitchen_total + consumable_total + monthly_total + capital_total + other_total

    # ── Build workbook ──
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"

    # Colors
    BLUE_DARK   = "1E3A5F"
    BLUE_MID    = "2563EB"
    BLUE_LIGHT  = "DBEAFE"
    GREY_ROW    = "F8FAFC"
    WHITE       = "FFFFFF"
    ORANGE      = "D97706"
    GREEN       = "16A34A"
    HEADER_BG   = "1E3A5F"

    def style_cell(cell, bold=False, bg=None, fg="000000", align="left", size=10, border=False, num_fmt=None):
        cell.font = Font(bold=bold, color=fg, size=size, name="Calibri")
        if bg:
            cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=False)
        if border:
            thin = Side(style="thin", color="D1D5DB")
            cell.border = Border(bottom=thin, top=thin, left=thin, right=thin)
        if num_fmt:
            cell.number_format = num_fmt

    # ── Title rows ──
    ws.merge_cells("A1:J1")
    ws["A1"] = "AB GROUP – EXPENSE SUMMARY"
    style_cell(ws["A1"], bold=True, bg=HEADER_BG, fg=WHITE, align="center", size=14)
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:J2")
    ws["A2"] = f"Month: {month_label}   |   Generated: {datetime.today().strftime('%d %b %Y')}"
    style_cell(ws["A2"], bg="2563EB", fg=WHITE, align="center", size=10)
    ws.row_dimensions[2].height = 18

    ws.row_dimensions[3].height = 8  # spacer

    # ── Column headers ──
    ws.row_dimensions[4].height = 22
    ws["A4"] = "Particulars"
    ws["B4"] = month_label
    ws["C4"] = "Total"
    for col in ["A4", "B4", "C4"]:
        style_cell(ws[col], bold=True, bg=HEADER_BG, fg=WHITE, align="center", size=10, border=True)

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16

    row = 5
    section_colors = {
        "Kitchen Consumables (a)": "1D4ED8",
        "Consumables (b)":         "0369A1",
        "Monthly Expenses ( C)":   "065F46",
        "Capital Expenses (d)":    "92400E",
    }
    section_subtotal_bgs = {
        "Kitchen Consumables (a)": "BFDBFE",
        "Consumables (b)":         "BAE6FD",
        "Monthly Expenses ( C)":   "A7F3D0",
        "Capital Expenses (d)":    "FEF3C7",
    }

    def write_section(section_name, heads, subtotal):
        nonlocal row
        # Section header
        ws.row_dimensions[row].height = 20
        ws[f"A{row}"] = section_name
        ws[f"B{row}"] = ""
        ws[f"C{row}"] = ""
        clr = section_colors.get(section_name, BLUE_DARK)
        for col in ["A", "B", "C"]:
            style_cell(ws[f"{col}{row}"], bold=True, bg=clr, fg=WHITE, align="left", size=10)
        row += 1

        # Item rows
        for i, head in enumerate(heads):
            ws.row_dimensions[row].height = 17
            amt = totals.get(head, 0)
            bg = WHITE if i % 2 == 0 else GREY_ROW
            ws[f"A{row}"] = f"  {head}"
            ws[f"B{row}"] = amt if amt else 0
            ws[f"C{row}"] = amt if amt else 0
            style_cell(ws[f"A{row}"], bg=bg, size=10)
            style_cell(ws[f"B{row}"], bg=bg, align="right", size=10, num_fmt='#,##0.00')
            style_cell(ws[f"C{row}"], bg=bg, align="right", size=10, num_fmt='#,##0.00')
            row += 1

        # Subtotal row
        ws.row_dimensions[row].height = 18
        st_bg = section_subtotal_bgs.get(section_name, BLUE_LIGHT)
        ws[f"A{row}"] = f"  Sub-total — {section_name}"
        ws[f"B{row}"] = subtotal
        ws[f"C{row}"] = subtotal
        style_cell(ws[f"A{row}"], bold=True, bg=st_bg, size=10)
        style_cell(ws[f"B{row}"], bold=True, bg=st_bg, align="right", size=10, num_fmt='#,##0.00')
        style_cell(ws[f"C{row}"], bold=True, bg=st_bg, align="right", size=10, num_fmt='#,##0.00')
        row += 1
        ws.row_dimensions[row].height = 5  # spacer
        row += 1

    write_section("Kitchen Consumables (a)", kitchen_heads, kitchen_total)
    write_section("Consumables (b)", consumable_heads, consumable_total)
    write_section("Monthly Expenses ( C)", monthly_heads, monthly_total)
    write_section("Capital Expenses (d)", capital_heads, capital_total)

    # Other / Uncategorised
    if other_total > 0:
        ws.row_dimensions[row].height = 20
        ws[f"A{row}"] = "Other / Uncategorised"
        ws[f"B{row}"] = ""
        ws[f"C{row}"] = ""
        for col in ["A", "B", "C"]:
            style_cell(ws[f"{col}{row}"], bold=True, bg="6B7280", fg=WHITE, size=10)
        row += 1
        other_items = df[df["Main Group"] == "Other"].groupby("Expense Head")["Amount"].sum()
        for i, (head, amt) in enumerate(other_items.items()):
            bg = WHITE if i % 2 == 0 else GREY_ROW
            ws[f"A{row}"] = f"  {head if head else '(Blank)'}"
            ws[f"B{row}"] = round(amt, 2)
            ws[f"C{row}"] = round(amt, 2)
            style_cell(ws[f"A{row}"], bg=bg, size=10)
            style_cell(ws[f"B{row}"], bg=bg, align="right", size=10, num_fmt='#,##0.00')
            style_cell(ws[f"C{row}"], bg=bg, align="right", size=10, num_fmt='#,##0.00')
            row += 1
        ws[f"A{row}"] = "  Sub-total — Other"
        ws[f"B{row}"] = other_total
        ws[f"C{row}"] = other_total
        style_cell(ws[f"A{row}"], bold=True, bg="E5E7EB", size=10)
        style_cell(ws[f"B{row}"], bold=True, bg="E5E7EB", align="right", size=10, num_fmt='#,##0.00')
        style_cell(ws[f"C{row}"], bold=True, bg="E5E7EB", align="right", size=10, num_fmt='#,##0.00')
        row += 1
        row += 1

    # ── Totals footer ──
    # Summary sub-totals section (matching the original)
    summaries = [
        ("Total Kitchen Consumables", kitchen_total, "BFDBFE"),
        ("Total Consumables",         consumable_total, "BAE6FD"),
        ("Total Monthly Expenses",    monthly_total,  "A7F3D0"),
        ("Total Capital Expenses",    capital_total,  "FEF3C7"),
        ("Other / Uncategorised",     other_total,    "E5E7EB"),
    ]
    ws[f"A{row}"] = "GRAND TOTAL"
    ws[f"B{row}"] = grand_total
    ws[f"C{row}"] = grand_total
    ws.row_dimensions[row].height = 22
    for col in ["A", "B", "C"]:
        style_cell(ws[f"{col}{row}"], bold=True, bg=HEADER_BG, fg=WHITE,
                   align="right" if col != "A" else "left", size=11,
                   border=True, num_fmt='#,##0.00' if col != "A" else None)
    row += 2

    ws[f"A{row}"] = "Breakdown"
    style_cell(ws[f"A{row}"], bold=True, bg="374151", fg=WHITE, size=10)
    style_cell(ws[f"B{row}"], bold=True, bg="374151", fg=WHITE, size=10)
    style_cell(ws[f"C{row}"], bold=True, bg="374151", fg=WHITE, size=10)
    row += 1
    for label, val, bg in summaries:
        ws[f"A{row}"] = f"  {label}"
        ws[f"B{row}"] = val
        ws[f"C{row}"] = val
        ws.row_dimensions[row].height = 17
        style_cell(ws[f"A{row}"], bg=bg, size=10)
        style_cell(ws[f"B{row}"], bg=bg, align="right", size=10, num_fmt='#,##0.00')
        style_cell(ws[f"C{row}"], bg=bg, align="right", size=10, num_fmt='#,##0.00')
        row += 1

    # Freeze top rows
    ws.freeze_panes = "A5"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─── Session State ─────────────────────────────────────────────────────────────
if "records" not in st.session_state:
    st.session_state.records = pd.DataFrame()
if "cash_entries" not in st.session_state:
    st.session_state.cash_entries = []
if "month_label" not in st.session_state:
    st.session_state.month_label = ""

# ─── Header ───────────────────────────────────────────────────────────────────
st.markdown("""
<div class="main-header">
    <h2 style="margin:0;font-size:1.5rem;">📊 AB Group — Expense Analyser</h2>
    <p style="margin:0.2rem 0 0;opacity:0.75;font-size:0.85rem;">Upload monthly expense sheet → Analyse → Download Summary Excel</p>
</div>
""", unsafe_allow_html=True)

# ─── Upload ───────────────────────────────────────────────────────────────────
uploaded = st.file_uploader(
    "📂 Upload Monthly Expense File (.xlsx)",
    type=["xlsx", "xls"],
    help="Upload the monthly UPI/cash expense file. F column (Expense Head) is used directly for categorisation."
)

if uploaded:
    try:
        df_parsed, sheet_name, month_label = parse_expense_sheet(uploaded)
        st.session_state.records    = df_parsed
        st.session_state.month_label = month_label
        n_uncat = df_parsed["Needs Category"].sum()
        st.success(f"✅ Loaded **{sheet_name}** — {len(df_parsed)} transactions · "
                   f"{'⚠️ ' + str(n_uncat) + ' missing Expense Head (added to Other)' if n_uncat else 'All categorised'}")
    except Exception as e:
        st.error(f"Error reading file: {e}")

# Combine records + cash
df = st.session_state.records.copy()
cash_df = pd.DataFrame(st.session_state.cash_entries) if st.session_state.cash_entries else pd.DataFrame()
if not cash_df.empty:
    combined = pd.concat([df, cash_df], ignore_index=True)
else:
    combined = df.copy()

# ─── PRIMARY: Download Summary Excel ──────────────────────────────────────────
if not combined.empty:
    st.markdown("---")
    col_dl1, col_dl2, col_dl3 = st.columns([2, 1, 2])
    with col_dl2:
        month_label = st.session_state.month_label or "Month"
        excel_buf = generate_summary_excel(combined, month_label)
        st.download_button(
            label="📥 Download Summary Excel",
            data=excel_buf,
            file_name=f"Summary_{month_label.replace(' ', '_').replace('/', '-')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary",
        )
    st.markdown("---")

# ─── Tabs ─────────────────────────────────────────────────────────────────────
tab1, tab2, tab3, tab4 = st.tabs(["📊 Analysis", "📋 All Records", "💵 Cash Entry", "🏷️ Categorise"])

# ══ TAB 1: ANALYSIS ══════════════════════════════════════════════════════════
with tab1:
    if combined.empty:
        st.info("👆 Upload a monthly expense file to see analysis.")
    else:
        total = combined["Amount"].sum()
        n_txn = len(combined)
        n_other = (combined["Main Group"] == "Other").sum()

        c1, c2, c3, c4 = st.columns(4)
        c1.metric("💰 Total Expenses",   fmt_inr(total))
        c2.metric("🧾 Transactions",      n_txn)
        c3.metric("📁 Month",             st.session_state.month_label or "—")
        c4.metric("⚠️ Uncategorised",     n_other, help="Added to 'Other' in output")

        st.divider()

        col1, col2 = st.columns(2)
        with col1:
            st.subheader("By Section")
            by_group = combined.groupby("Main Group")["Amount"].sum().reset_index().sort_values("Amount", ascending=False)
            fig = px.pie(by_group, values="Amount", names="Main Group", hole=0.4,
                         color_discrete_sequence=px.colors.qualitative.Bold)
            fig.update_traces(textposition="inside", textinfo="percent+label")
            fig.update_layout(margin=dict(t=10,b=10,l=10,r=10), showlegend=False, height=300)
            st.plotly_chart(fig, use_container_width=True)

        with col2:
            st.subheader("Top 12 Expense Heads")
            by_head = combined.groupby("Expense Head")["Amount"].sum().reset_index()
            by_head = by_head[by_head["Expense Head"].str.strip() != ""].sort_values("Amount", ascending=False)
            fig2 = px.bar(by_head.head(12), x="Amount", y="Expense Head", orientation="h",
                          color="Amount", color_continuous_scale="Blues",
                          text=by_head.head(12)["Amount"].apply(fmt_inr))
            fig2.update_traces(textposition="outside")
            fig2.update_layout(margin=dict(t=10,b=10,l=10,r=60),
                               yaxis=dict(autorange="reversed"),
                               coloraxis_showscale=False, height=300,
                               xaxis_title="", yaxis_title="")
            st.plotly_chart(fig2, use_container_width=True)

        # Section breakdown table
        st.subheader("📋 Section Breakdown")
        rows = []
        for section, heads in SUMMARY_STRUCTURE.items():
            amt = sum(combined[combined["Expense Head"].str.strip().str.lower() == h.strip().lower()]["Amount"].sum()
                      for h in heads)
            rows.append({"Section": section, "Amount (₹)": f"₹{amt:,.2f}",
                         "% of Total": f"{(amt/total*100):.1f}%" if total else "0%"})
        other_amt = combined[combined["Main Group"] == "Other"]["Amount"].sum()
        if other_amt > 0:
            rows.append({"Section": "Other / Uncategorised", "Amount (₹)": f"₹{other_amt:,.2f}",
                         "% of Total": f"{(other_amt/total*100):.1f}%"})
        st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

# ══ TAB 2: ALL RECORDS ═══════════════════════════════════════════════════════
with tab2:
    if combined.empty:
        st.info("No records yet.")
    else:
        col_s, col_f = st.columns([2, 2])
        with col_s:
            search = st.text_input("🔍 Search description", placeholder="salary, fuel…")
        with col_f:
            grp_filter = st.multiselect("Filter by Section", options=sorted(combined["Main Group"].unique()))

        view = combined.copy()
        if search:
            view = view[view["Description"].str.contains(search, case=False, na=False)]
        if grp_filter:
            view = view[view["Main Group"].isin(grp_filter)]

        st.caption(f"**{len(view)}** records · Total: ₹{view['Amount'].sum():,.2f}")
        show = ["Date", "Description", "Amount", "Expense Head", "Main Group", "Account"]
        show = [c for c in show if c in view.columns]
        st.dataframe(
            view[show].sort_values("Amount", ascending=False),
            use_container_width=True, hide_index=True,
            column_config={"Amount": st.column_config.NumberColumn("Amount (₹)", format="₹%.2f")}
        )

# ══ TAB 3: CASH ENTRY ════════════════════════════════════════════════════════
with tab3:
    col1, col2 = st.columns(2, gap="large")
    with col1:
        st.subheader("➕ Add Cash Expense")
        with st.form("cash_form", clear_on_submit=True):
            title  = st.text_input("Title / Description *", placeholder="e.g. Petrol for generator")
            amount = st.number_input("Amount (₹) *", min_value=0.0, step=10.0, format="%.2f")
            head   = st.selectbox("Expense Head *", ["— Select —"] + ALL_KNOWN_HEADS)
            custom = st.text_input("Or type custom Expense Head", placeholder="Leave blank if selected above")
            if st.form_submit_button("✅ Add", use_container_width=True, type="primary"):
                final_head = custom.strip() if custom.strip() else (head if head != "— Select —" else "")
                if not title:            st.error("Title required.")
                elif amount <= 0:        st.error("Amount must be > 0.")
                elif not final_head:     st.error("Select or type an expense head.")
                else:
                    st.session_state.cash_entries.append({
                        "Date": datetime.today().strftime("%d/%m/%y"),
                        "Description": title,
                        "Amount": float(amount),
                        "Expense Head": final_head,
                        "Main Group": get_group(final_head),
                        "Account": "Cash",
                        "Needs Category": False,
                    })
                    st.success(f"Added ₹{amount:,.0f} — {final_head}")
                    st.rerun()

    with col2:
        st.subheader(f"💵 Cash Entries ({len(st.session_state.cash_entries)})")
        if not st.session_state.cash_entries:
            st.info("No cash entries yet.")
        else:
            cdf = pd.DataFrame(st.session_state.cash_entries)[["Date","Description","Amount","Expense Head","Main Group"]]
            st.dataframe(cdf, use_container_width=True, hide_index=True,
                         column_config={"Amount": st.column_config.NumberColumn("₹", format="₹%.2f")})
            st.metric("Total Cash", f"₹{cdf['Amount'].sum():,.2f}")
            if st.button("🗑️ Clear all cash entries"):
                st.session_state.cash_entries = []
                st.rerun()

# ══ TAB 4: CATEGORISE ════════════════════════════════════════════════════════
with tab4:
    st.subheader("🏷️ Entries with Missing Expense Head")
    if df.empty:
        st.info("Upload a file first.")
    else:
        missing = df[df["Needs Category"] == True].copy()
        if missing.empty:
            st.success("✅ All entries have an Expense Head.")
        else:
            st.warning(f"{len(missing)} entries have no Expense Head — currently assigned to **Other**.")
            for idx, item in missing.iterrows():
                c1, c2, c3, c4 = st.columns([3, 2, 2, 1])
                c1.caption(f"**{item['Description'][:55]}**  \n{item['Date']} · ₹{item['Amount']:,.0f}")
                new_head = c2.selectbox("Expense Head", ["— Select —"] + ALL_KNOWN_HEADS, key=f"h_{idx}")
                custom_h = c3.text_input("Custom", key=f"ch_{idx}", placeholder="or type here")
                c4.write("")
                if c4.button("Save", key=f"s_{idx}"):
                    final = custom_h.strip() if custom_h.strip() else (new_head if new_head != "— Select —" else "")
                    if final:
                        st.session_state.records.at[idx, "Expense Head"] = final
                        st.session_state.records.at[idx, "Main Group"]   = get_group(final)
                        st.session_state.records.at[idx, "Needs Category"] = False
                        st.rerun()

